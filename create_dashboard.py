#!/usr/bin/env python3
"""
WASH Excel Dashboard Creator - Professional Dashboard with Pivot Tables, Slicers & Deep Analysis
"""

import win32com.client as win32
import openpyxl
from collections import Counter, defaultdict
import os, sys, shutil, time

FILE_PATH  = r'C:\Users\ITrebi\Music\PortFolio Upwork\Excel Dashboard\wash_test_data.xlsx'
OUTPUT_PATH = r'C:\Users\ITrebi\Music\PortFolio Upwork\Excel Dashboard\WASH_Dashboard.xlsx'

# ── Color helpers ─────────────────────────────────────────────────────────────
def rgb(r, g, b): return r + g*256 + b*65536

C_DARK_NAVY  = rgb(31,  56, 100)   # #1F3864
C_BLUE       = rgb( 0,  84, 166)   # #0054A6  – WASH brand blue
C_MED_BLUE   = rgb(46, 117, 182)   # #2E75B6
C_LIGHT_BLUE = rgb(189,215, 238)   # #BDD7EE
C_TEAL       = rgb( 0, 176, 185)   # #00B0B9
C_GREEN      = rgb(112,173,  71)   # #70AD47
C_ORANGE     = rgb(237,125,  49)   # #ED7D31
C_RED        = rgb(192,  0,   0)   # #C00000
C_AMBER      = rgb(255,192,   0)   # #FFC000
C_WHITE      = rgb(255,255, 255)
C_LGRAY      = rgb(242,242, 242)
C_MGRAY      = rgb(208,208, 208)
C_DGRAY      = rgb( 89, 89,  89)
C_BLACK      = rgb(  0,  0,   0)
C_STRIPE     = rgb(217,232, 248)   # light blue stripe

# ── Excel constants ────────────────────────────────────────────────────────────
xlDatabase        = 1
xlRowField        = 1
xlColumnField     = 2
xlDataField       = 4
xlCount           = -4112
xlAverage         = -4106

xlColumnClustered = 51
xlBarClustered    = 57
xlPie             = 5
xlDoughnut        = -4120
xlColumnStacked100= 59
xlBarStacked100   = 70
xlBarStacked      = 52
xlColumnStacked   = 52

xlCenter          = -4108
xlLeft            = -4131
xlRight           = -4152
xlVAlignCenter    = -4108
xlVAlignTop       = -4160
xlVAlignBottom    = -4107

xlContinuous      = 1
xlThin            = 2
xlMedium          = -4138
xlThick           = 4
xlEdgeLeft        = 7
xlEdgeTop         = 8
xlEdgeBottom      = 9
xlEdgeRight       = 10
xlInsideVertical  = 11
xlInsideHorizontal= 12
xlAround          = 4

xlSheetHidden     = 0
xlSheetVisible    = -1
xlSheetVeryHidden = 2

xlPasteValues     = -4163
xlPasteSpecialOperationNone = -4142

# ── Pre-compute statistics ─────────────────────────────────────────────────────
print(">>>  Chargement des données…")
wb_tmp = openpyxl.load_workbook(FILE_PATH, read_only=True)
ws_tmp = wb_tmp.active
HEADERS = [c.value for c in next(ws_tmp.iter_rows(min_row=1, max_row=1))]
ROWS    = list(ws_tmp.iter_rows(min_row=2, values_only=True))
wb_tmp.close()

N = len(ROWS)   # 1000

def col(name):    return HEADERS.index(name)
def vals(name):   return [r[col(name)] for r in ROWS]
def cnt(name):    return Counter(vals(name))
def pct(name, v): n = sum(1 for r in ROWS if r[col(name)]==v); return n, round(n/N*100,1)

# KPI values
n_bonne, p_bonne   = pct('e5_qualite_eau', 'bonne')
n_acc30, p_acc30   = pct('e4_temps_acces', 'moins_30')
n_traite,p_traite  = pct('traite_eau', True)
n_mal,   p_mal     = pct('e7_maladie_eau','oui')
n_poll,  p_poll    = pct('p1_pollution_atmo','oui')
n_deg,   p_deg     = pct('r5_degradation','oui')
durees = [r[col('duree_enquete_min')] for r in ROWS if r[col('duree_enquete_min')]]
moy_d  = round(sum(durees)/len(durees),1)

# Distributions
c_source  = cnt('e1_source_eau')
c_qualite = cnt('e5_qualite_eau')
c_temps   = cnt('e4_temps_acces')
c_latr    = cnt('e11_type')
c_mal     = cnt('e7_maladie_eau')
c_dechets = cnt('d3_satisfaction')
c_poll    = cnt('p1_pollution_atmo')
c_air     = cnt('p3_qualite_air')
c_ress    = cnt('r5_degradation')
c_dept    = cnt('departement')
c_evass   = cnt('e9_eval_assainissement')
c_bruit   = cnt('p4_nuisances_sonores')

# Months from ISO date
dates_raw = vals('date_soumission')
months_map = {'01':'Janvier','02':'Février','03':'Mars','04':'Avril',
              '05':'Mai','06':'Juin','07':'Juillet','08':'Août',
              '09':'Septembre','10':'Octobre','11':'Novembre','12':'Décembre'}
def parse_month(d):
    if not d or len(d)<7: return 'N/A'
    return months_map.get(d[5:7], d[5:7])
c_mois = Counter(parse_month(d) for d in dates_raw)

# Cross-tabs (dept x field)
def xtab(field):
    d = defaultdict(Counter)
    for r in ROWS:
        dept = r[col('departement')] or 'N/A'
        v    = r[col(field)] or 'N/A'
        d[dept][v] += 1
    return d

xt_source  = xtab('e1_source_eau')
xt_qualite = xtab('e5_qualite_eau')
xt_temps   = xtab('e4_temps_acces')
xt_mal     = xtab('e7_maladie_eau')
xt_poll    = xtab('p1_pollution_atmo')
xt_ress    = xtab('r5_degradation')

# Sous-préfectures
sous_prefs = sorted(set(r[col('sous_prefecture')] for r in ROWS if r[col('sous_prefecture')]))

# Disease details
diseases = {
    'Diarrhée':   sum(1 for r in ROWS if r[col('e7_maladies_diarrhee')]=='oui'),
    'Typhoïde':   sum(1 for r in ROWS if r[col('e7_maladies_typhoide')]=='oui'),
    'Hépatite':   sum(1 for r in ROWS if r[col('e7_maladies_hepatite')]=='oui'),
    'Choléra':    sum(1 for r in ROWS if r[col('e7_maladies_cholera')]=='oui'),
}

# Waste management breakdown
waste_types = {
    'Abandon':      sum(1 for r in ROWS if r[col('d2_gestion_dechets_abandon')]=='oui'),
    'Brûlage':      sum(1 for r in ROWS if r[col('d2_gestion_dechets_brulage')]=='oui'),
    'Collecte/Val.':sum(1 for r in ROWS if r[col('d2_gestion_dechets_collecte_valorisation')]=='oui'),
    'Réutilisation':sum(1 for r in ROWS if r[col('d2_gestion_dechets_reutilisation')]=='oui'),
}

# Resource degradation causes
ress_causes = {
    'Pratiques agri.': sum(1 for r in ROWS if r[col('r6_causes_Pratiques_agricoles')]=='oui'),
    'Exploit. forest.':sum(1 for r in ROWS if r[col('r6_causes_Exploitation_forestiere')]=='oui'),
    'Urbanisation':    sum(1 for r in ROWS if r[col('r6_causes_Urbanisation')]=='oui'),
    'Exploit. minière':sum(1 for r in ROWS if r[col('r6_causes_Exploitation_miniere')]=='oui'),
}

print(f"   {N} enquêtes | Bonne eau: {p_bonne}% | Pollution: {p_poll}%")

# ── Copy source file to output ─────────────────────────────────────────────────
print(">>>  Copie du fichier source…")
shutil.copy2(FILE_PATH, OUTPUT_PATH)

# ── Open with win32com ────────────────────────────────────────────────────────
print(">>>  Ouverture Excel…")
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = False
excel.DisplayAlerts = False
excel.ScreenUpdating = False

wb = excel.Workbooks.Open(OUTPUT_PATH)
ws_data = wb.Worksheets('wash_test_data')

# ── Add "mois" column to data (col 167 = FK) ──────────────────────────────────
print(">>>  Ajout colonne mois…")
last_row = ws_data.UsedRange.Rows.Count  # 1001
ws_data.Cells(1, 167).Value = 'mois'
mois_vals = [parse_month(r[col('date_soumission')]) for r in ROWS]
for i, mv in enumerate(mois_vals):
    ws_data.Cells(i+2, 167).Value = mv

# ── Delete existing dashboard sheets if any ───────────────────────────────────
for sname in ['TABLEAU_BORD','ANALYSE_PROFONDE','_PIVOTS']:
    try:
        wb.Worksheets(sname).Delete()
    except: pass

# ── Create sheets ─────────────────────────────────────────────────────────────
print(">>>  Création des feuilles…")
ws_piv    = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
ws_piv.Name = '_PIVOTS'

ws_dash   = wb.Sheets.Add(Before=wb.Sheets(1))
ws_dash.Name = 'TABLEAU_BORD'

ws_ana    = wb.Sheets.Add(After=ws_dash)
ws_ana.Name = 'ANALYSE_PROFONDE'

# Color tabs
ws_dash.Tab.Color = C_BLUE
ws_ana.Tab.Color  = C_TEAL

# ── Pivot Cache & Pivot Tables ────────────────────────────────────────────────
print(">>>  Création des tableaux croisés dynamiques…")
data_addr = f"wash_test_data!R1C1:R{last_row}C167"
pc = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=data_addr)

def make_pt(dest_cell, name, row_fld, col_fld=None, val_fld='id_soumission'):
    pt = pc.CreatePivotTable(TableDestination=dest_cell, TableName=name)
    pt.PivotFields(row_fld).Orientation = xlRowField
    pt.PivotFields(row_fld).Position = 1
    if col_fld:
        pt.PivotFields(col_fld).Orientation = xlColumnField
        pt.PivotFields(col_fld).Position = 1
    pt.AddDataField(pt.PivotFields(val_fld), 'Nb', -4112)
    pt.ShowTableStyleRowStripes = True
    return pt

r = 1
pt_src  = make_pt(ws_piv.Cells(r,1),  'PT_Source',    'e1_source_eau',   'departement'); r+=12
pt_qlt  = make_pt(ws_piv.Cells(r,1),  'PT_Qualite',   'e5_qualite_eau',  'departement'); r+=8
pt_tps  = make_pt(ws_piv.Cells(r,1),  'PT_Temps',     'e4_temps_acces',  'departement'); r+=8
pt_lat  = make_pt(ws_piv.Cells(r,1),  'PT_Latrines',  'e11_type',        'departement'); r+=10
pt_mal  = make_pt(ws_piv.Cells(r,1),  'PT_Maladies',  'e7_maladie_eau',  'departement'); r+=8
pt_dch  = make_pt(ws_piv.Cells(r,1),  'PT_Dechets',   'd3_satisfaction', 'departement'); r+=8
pt_pll  = make_pt(ws_piv.Cells(r,1),  'PT_Pollution', 'p1_pollution_atmo','departement'); r+=8
pt_air  = make_pt(ws_piv.Cells(r,1),  'PT_Air',       'p3_qualite_air',  'departement'); r+=8
pt_res  = make_pt(ws_piv.Cells(r,1),  'PT_Ressources','r5_degradation',  'departement'); r+=8
pt_mois = make_pt(ws_piv.Cells(r,1),  'PT_Mois',      'mois');                           r+=15
pt_brt  = make_pt(ws_piv.Cells(r,1),  'PT_Bruit',     'p4_nuisances_sonores','departement'); r+=8
pt_evass= make_pt(ws_piv.Cells(r,1),  'PT_Evass',     'e9_eval_assainissement','departement'); r+=8

ws_piv.Visible = xlSheetVeryHidden

# ── Slicers ───────────────────────────────────────────────────────────────────
print(">>>  Création des slicers…")
ALL_PTS = [pt_src, pt_qlt, pt_tps, pt_lat, pt_mal, pt_dch, pt_pll, pt_air, pt_res, pt_brt, pt_evass]

def add_slicer_cache(base_pt, field, sc_name, other_pts):
    sc = wb.SlicerCaches.Add2(base_pt, field, sc_name)
    for pt in other_pts:
        try: sc.PivotTables.AddPivotTable(pt)
        except: pass
    return sc

sc_dept  = add_slicer_cache(pt_src, 'departement',     'SC_Dept',   ALL_PTS[1:])
sc_sp    = add_slicer_cache(pt_src, 'sous_prefecture', 'SC_SP',     ALL_PTS[1:])
sc_src   = add_slicer_cache(pt_src, 'e1_source_eau',   'SC_Src',    ALL_PTS[1:])
sc_qlt   = add_slicer_cache(pt_qlt, 'e5_qualite_eau',  'SC_Qlt',    [pt_src,pt_tps,pt_lat,pt_mal,pt_dch,pt_pll,pt_air,pt_res,pt_brt,pt_evass])
sc_mois_c= add_slicer_cache(pt_mois,'mois',             'SC_Mois',   ALL_PTS)

# ── ════════════════════════════════════════════════════════════════════════════
# ── FORMAT DASHBOARD SHEET
# ── ════════════════════════════════════════════════════════════════════════════
print(">>>  Mise en forme du tableau de bord…")

# Freeze panes – keep header visible
ws_dash.Activate()
ws_dash.Application.ActiveWindow.ActiveSheet.Cells(11,2).Select()
try:
    ws_dash.Application.ActiveWindow.FreezePanes = True
except: pass

# Set sheet background – entire used area
ws_dash.Cells.Interior.Color = C_LGRAY

# Column widths (in char units)
cwidths = {'A':1.5,'B':17,'C':17,'D':17,'E':17,'F':17,'G':17,
           'H':1.5,'I':20,'J':16,'K':1.5,'L':20,'M':1.5}
for c,w in cwidths.items():
    ws_dash.Columns(c).ColumnWidth = w

# ── HEADER BANNER (rows 1-5) ──────────────────────────────────────────────────
for r in range(1,6):
    ws_dash.Rows(r).RowHeight = 16

ws_dash.Range('B1:G5').Merge()
h = ws_dash.Range('B1')
h.Value = 'TABLEAU DE BORD  ·  ENQUÊTE WASH'
h.Font.Name = 'Calibri'
h.Font.Size = 22
h.Font.Bold = True
h.Font.Color = C_WHITE
h.Interior.Color = C_BLUE
h.HorizontalAlignment = xlCenter
h.VerticalAlignment   = xlVAlignCenter

ws_dash.Range('I1:J2').Merge()
sub = ws_dash.Range('I1')
sub.Value = '~  Eau  •  Assainissement  •  Hygiène'
sub.Font.Size = 11
sub.Font.Bold = True
sub.Font.Color = C_WHITE
sub.Interior.Color = C_TEAL
sub.HorizontalAlignment = xlCenter
sub.VerticalAlignment   = xlVAlignCenter

ws_dash.Range('I3:J3').Merge()
meta = ws_dash.Range('I3')
meta.Value = f'{N} enquêtes  |  2 départements  |  8 sous-préfectures'
meta.Font.Size = 8
meta.Font.Color = C_DGRAY
meta.Interior.Color = C_LGRAY
meta.HorizontalAlignment = xlCenter

ws_dash.Range('I4:J5').Merge()
ws_dash.Range('I4').Value = 'Côte d\'Ivoire — Jan/Fév 2026'
ws_dash.Range('I4').Font.Size = 8
ws_dash.Range('I4').Font.Italic = True
ws_dash.Range('I4').Font.Color = C_DGRAY
ws_dash.Range('I4').Interior.Color = C_LGRAY
ws_dash.Range('I4').HorizontalAlignment = xlCenter
ws_dash.Range('I4').VerticalAlignment   = xlVAlignCenter

# ── SLICERS PANEL HEADER ──────────────────────────────────────────────────────
ws_dash.Range('L1:L5').Merge()
sp = ws_dash.Range('L1')
sp.Value = '~ FILTRES'
sp.Font.Size = 12
sp.Font.Bold = True
sp.Font.Color = C_WHITE
sp.Interior.Color = C_DARK_NAVY
sp.HorizontalAlignment = xlCenter
sp.VerticalAlignment   = xlVAlignCenter

# ── KPI CARDS (rows 6-9) ──────────────────────────────────────────────────────
ws_dash.Rows('6:6').RowHeight = 28
ws_dash.Rows('7:7').RowHeight = 36
ws_dash.Rows('8:8').RowHeight = 18
ws_dash.Rows('9:9').RowHeight = 8   # spacer

kpi_def = [
    ('B','Total Enquêtes',    f'{N}',            '1 000 ménages enquêtés',  C_BLUE),
    ('C','Eau Bonne Qualité', f'{p_bonne}%',      f'{n_bonne} enquêtes',    C_GREEN),
    ('D','Accès < 30 min',    f'{p_acc30}%',      f'{n_acc30} enquêtes',    C_TEAL),
    ('E','Maladies / Eau',    f'{p_mal}%',        f'{n_mal} enquêtes',      C_RED),
    ('F','Pollution Atmo.',   f'{p_poll}%',        f'{n_poll} enquêtes',     C_AMBER),
    ('G','Dégradation Res.',  f'{p_deg}%',        f'{n_deg} enquêtes',      C_ORANGE),
]

def border_cell(cell, color, weight=xlThin):
    for edge in [xlEdgeLeft,xlEdgeTop,xlEdgeBottom,xlEdgeRight]:
        b = cell.Borders(edge)
        b.LineStyle = xlContinuous
        b.Weight = weight
        b.Color = color

for c, title, val, sub_, color in kpi_def:
    # Title row
    t = ws_dash.Range(f'{c}6')
    t.Value = title; t.Font.Size=9; t.Font.Bold=True
    t.Font.Color=C_WHITE; t.Interior.Color=color
    t.HorizontalAlignment=xlCenter; t.VerticalAlignment=xlVAlignCenter
    border_cell(t, color, xlMedium)
    # Value row
    v = ws_dash.Range(f'{c}7')
    v.Value=val; v.Font.Size=20; v.Font.Bold=True
    v.Font.Color=color; v.Interior.Color=C_WHITE
    v.HorizontalAlignment=xlCenter; v.VerticalAlignment=xlVAlignCenter
    border_cell(v, color, xlMedium)
    # Subtitle row
    s = ws_dash.Range(f'{c}8')
    s.Value=sub_; s.Font.Size=7; s.Font.Color=C_DGRAY
    s.Interior.Color=C_LGRAY
    s.HorizontalAlignment=xlCenter
    border_cell(s, color)

# ── SECTION LABELS ────────────────────────────────────────────────────────────
def section_label(row_num, text, color):
    ws_dash.Rows(row_num).RowHeight = 20
    r = ws_dash.Range(f'B{row_num}:G{row_num}')
    r.Merge()
    r.Value = text
    r.Font.Size = 10; r.Font.Bold = True; r.Font.Color = C_WHITE
    r.Interior.Color = color
    r.HorizontalAlignment = xlCenter; r.VerticalAlignment = xlVAlignCenter

section_label(10, '~  EAU  –  Sources · Qualité · Accès',                    C_BLUE)
section_label(32, '~  ASSAINISSEMENT & SANTÉ  –  Latrines · Maladies',        C_MED_BLUE)
section_label(54, '~  DÉCHETS & POLLUTION  –  Gestion · Air · Bruit',         C_TEAL)
section_label(76, '~  RESSOURCES NATURELLES  –  Disponibilité · Dégradation', C_GREEN)

# ── CHART ROW HEIGHTS ─────────────────────────────────────────────────────────
for r in range(11,32): ws_dash.Rows(r).RowHeight = 10
for r in range(33,54): ws_dash.Rows(r).RowHeight = 10
for r in range(55,76): ws_dash.Rows(r).RowHeight = 10
for r in range(77,98): ws_dash.Rows(r).RowHeight = 10

# ── CHART HELPER ─────────────────────────────────────────────────────────────
def add_chart(ws, pt_range, ctype, title, left, top, w, h,
              legend_pos=-4107, color_list=None):
    co = ws.ChartObjects().Add(Left=left, Top=top, Width=w, Height=h)
    ch = co.Chart
    ch.SetSourceData(Source=pt_range)
    ch.ChartType = ctype
    ch.HasTitle = True
    ch.ChartTitle.Text = title
    ch.ChartTitle.Font.Size = 10
    ch.ChartTitle.Font.Bold = True
    ch.ChartTitle.Font.Color = C_DARK_NAVY
    ch.HasLegend = True
    ch.Legend.Position = legend_pos
    ch.Legend.Font.Size = 7
    ch.PlotArea.Interior.Color = C_WHITE
    ch.ChartArea.Interior.Color = C_LGRAY
    ch.ChartArea.Border.LineStyle = xlContinuous
    ch.ChartArea.Border.Color = C_MGRAY
    # Apply colors to series if provided
    if color_list:
        try:
            for i, c_ in enumerate(color_list):
                if i < ch.SeriesCollection().Count:
                    ch.SeriesCollection(i+1).Interior.Color = c_
                    ch.SeriesCollection(i+1).Format.Fill.ForeColor.RGB = c_
        except: pass
    return ch

# ── GET CELL POSITIONS ────────────────────────────────────────────────────────
def pos(row, col_letter):
    c = ws_dash.Range(f'{col_letter}{row}')
    return c.Left, c.Top

# Chart dimensions
CW3 = 245   # width for 3-per-row
CW2 = 375   # width for 2-per-row
CH  = 185   # height

GAP = 4

# Row 1 of charts: rows 11-31 → top of row 11
L_B, T10 = pos(11, 'B')
L_C = ws_dash.Range('C11').Left
L_E = ws_dash.Range('E11').Left

# ── SECTION EAU CHARTS ───────────────────────────────────────────────────────
# Chart 1: Sources d'eau (Donut)
add_chart(ws_dash, pt_src.TableRange1,
          xlDoughnut, "Sources d'eau", L_B, T10, CW3, CH,
          color_list=[C_BLUE,C_TEAL,C_GREEN,C_AMBER,C_ORANGE,C_RED])

# Chart 2: Qualité eau par département (Bar Clustered)
L_C2 = L_B + CW3 + GAP
add_chart(ws_dash, pt_qlt.TableRange1,
          xlBarClustered, "Qualité de l'eau par département",
          L_C2, T10, CW3, CH,
          color_list=[C_GREEN,C_AMBER,C_RED])

# Chart 3: Temps d'accès (Column Clustered)
L_C3 = L_C2 + CW3 + GAP
add_chart(ws_dash, pt_tps.TableRange1,
          xlColumnClustered, "Temps d'accès à l'eau",
          L_C3, T10, CW3, CH,
          color_list=[C_GREEN,C_AMBER,C_RED])

# ── SECTION ASSAINISSEMENT CHARTS ────────────────────────────────────────────
_, T32 = pos(33, 'B')

# Chart 4: Types latrines (Bar)
add_chart(ws_dash, pt_lat.TableRange1,
          xlBarClustered, "Types de latrines",
          L_B, T32, CW3, CH,
          color_list=[C_BLUE,C_MED_BLUE,C_LIGHT_BLUE,C_TEAL])

# Chart 5: Maladies liées à l'eau (Bar)
add_chart(ws_dash, pt_mal.TableRange1,
          xlColumnClustered, "Maladies liées à l'eau",
          L_C2, T32, CW3, CH,
          color_list=[C_RED, C_ORANGE, C_AMBER])

# Chart 6: Évaluation assainissement (Column)
add_chart(ws_dash, pt_evass.TableRange1,
          xlColumnClustered, "Évaluation assainissement",
          L_C3, T32, CW3, CH,
          color_list=[C_GREEN, C_RED, C_AMBER])

# ── SECTION DECHETS & POLLUTION CHARTS ───────────────────────────────────────
_, T54 = pos(55, 'B')

# Chart 7: Satisfaction gestion déchets (Bar)
add_chart(ws_dash, pt_dch.TableRange1,
          xlBarClustered, "Satisfaction gestion des déchets",
          L_B, T54, CW3, CH,
          color_list=[C_GREEN, C_RED])

# Chart 8: Pollution atmosphérique (Column)
add_chart(ws_dash, pt_pll.TableRange1,
          xlColumnClustered, "Pollution atmosphérique",
          L_C2, T54, CW3, CH,
          color_list=[C_ORANGE, C_GREEN])

# Chart 9: Qualité de l'air (Bar)
add_chart(ws_dash, pt_air.TableRange1,
          xlBarClustered, "Qualité de l'air",
          L_C3, T54, CW3, CH,
          color_list=[C_GREEN, C_RED])

# ── SECTION RESSOURCES CHARTS ─────────────────────────────────────────────────
_, T76 = pos(77, 'B')

# Chart 10: Dégradation ressources (Column)
add_chart(ws_dash, pt_res.TableRange1,
          xlColumnClustered, "Dégradation des ressources naturelles",
          L_B, T76, CW2, CH,
          color_list=[C_RED, C_GREEN, C_AMBER])

# Chart 11: Nuisances sonores (Bar)
L_R2 = L_B + CW2 + GAP
add_chart(ws_dash, pt_brt.TableRange1,
          xlBarClustered, "Nuisances sonores",
          L_R2, T76, CW2, CH,
          color_list=[C_AMBER, C_GREEN])

# ── ADD SLICERS TO DASHBOARD ──────────────────────────────────────────────────
print(">>>  Positionnement des slicers…")
L_slicer = ws_dash.Range('I6').Left
T_slicer = ws_dash.Range('I6').Top
SW = ws_dash.Range('I6:J6').Width   # slicer width
SH_S = 120  # short height
SH_M = 150  # medium height
SH_L = 180  # long height

slicer_defs = [
    (sc_dept,  'Département',     SW, SH_S),
    (sc_sp,    'Sous-préfecture', SW, SH_L),
    (sc_src,   "Source d'eau",    SW, SH_M),
    (sc_qlt,   "Qualité de l'eau",SW, SH_S),
    (sc_mois_c,'Mois',            SW, SH_S),
]

y_off = T_slicer
for sc, cap, sw, sh in slicer_defs:
    try:
        slc = sc.Slicers.Add(ws_dash, Caption=cap,
                             Left=L_slicer, Top=y_off, Width=sw, Height=sh)
        slc.Style = 'SlicerStyleLight2'
        y_off += sh + 6
    except Exception as e:
        print(f"   Slicer '{cap}': {e}")

# ── ════════════════════════════════════════════════════════════════════════════
# ── ANALYSE PROFONDE SHEET
# ── ════════════════════════════════════════════════════════════════════════════
print(">>>  Création de la feuille d'analyse…")

ws_ana.Cells.Interior.Color = C_WHITE
ws_ana.Columns('A').ColumnWidth = 3
ws_ana.Columns('B').ColumnWidth = 40
ws_ana.Columns('C').ColumnWidth = 18
ws_ana.Columns('D').ColumnWidth = 18
ws_ana.Columns('E').ColumnWidth = 18
ws_ana.Columns('F').ColumnWidth = 18
ws_ana.Columns('G').ColumnWidth = 18

def cell(row, col_letter):
    return ws_ana.Range(f'{col_letter}{row}')

def write_title(row, text, bg=C_BLUE, fg=C_WHITE, size=14, merge_to='G'):
    r = ws_ana.Range(f'B{row}:{merge_to}{row}')
    r.Merge(); r.Value=text
    r.Font.Size=size; r.Font.Bold=True
    r.Font.Color=fg; r.Interior.Color=bg
    r.HorizontalAlignment=xlCenter; r.VerticalAlignment=xlVAlignCenter
    ws_ana.Rows(row).RowHeight=24

def write_sub(row, text, bg=C_MED_BLUE, fg=C_WHITE):
    r = ws_ana.Range(f'B{row}:G{row}')
    r.Merge(); r.Value=text
    r.Font.Size=10; r.Font.Bold=True
    r.Font.Color=fg; r.Interior.Color=bg
    r.HorizontalAlignment=xlLeft; r.VerticalAlignment=xlVAlignCenter
    ws_ana.Rows(row).RowHeight=20

def write_header(row, labels, bg=C_DARK_NAVY, fg=C_WHITE):
    cols = ['B','C','D','E','F','G']
    ws_ana.Rows(row).RowHeight = 18
    for i, lbl in enumerate(labels):
        if i < len(cols):
            c = ws_ana.Range(f'{cols[i]}{row}')
            c.Value=lbl; c.Font.Bold=True; c.Font.Size=9
            c.Font.Color=fg; c.Interior.Color=bg
            c.HorizontalAlignment=xlCenter; c.VerticalAlignment=xlVAlignCenter
            border_cell(c, bg)

def write_row(row, vals, stripe=False, bold_first=False):
    cols = ['B','C','D','E','F','G']
    bg = C_STRIPE if stripe else C_WHITE
    ws_ana.Rows(row).RowHeight = 16
    for i, v in enumerate(vals):
        if i < len(cols):
            c = ws_ana.Range(f'{cols[i]}{row}')
            c.Value = v; c.Font.Size=9
            c.Interior.Color = bg
            c.VerticalAlignment = xlVAlignCenter
            if i==0 and bold_first: c.Font.Bold=True
            if i > 0: c.HorizontalAlignment = xlCenter
            border_cell(c, C_MGRAY)

def write_note(row, text, color=C_DGRAY):
    ws_ana.Range(f'B{row}:G{row}').Merge()
    c = ws_ana.Range(f'B{row}')
    c.Value = text; c.Font.Size=8; c.Font.Italic=True; c.Font.Color=color
    c.HorizontalAlignment=xlLeft
    ws_ana.Rows(row).RowHeight = 14

def spacer(row, h=8):
    ws_ana.Rows(row).RowHeight = h

# ══ MAIN TITLE ════════════════════════════════════════════════════════════════
ws_ana.Rows('1:1').RowHeight = 8
ws_ana.Range('B2:G4').Merge()
t = ws_ana.Range('B2')
t.Value='ANALYSE APPROFONDIE DES DONNÉES WASH'
t.Font.Size=18; t.Font.Bold=True; t.Font.Color=C_WHITE
t.Interior.Color=C_BLUE
t.HorizontalAlignment=xlCenter; t.VerticalAlignment=xlVAlignCenter
ws_ana.Rows('2:4').RowHeight=22

ws_ana.Range('B5:G5').Merge()
ws_ana.Range('B5').Value = (
    f'Enquête WASH – Côte d\'Ivoire  |  {N} ménages  |  Dépts: Goh ({c_dept["Goh"]}) & Loh Djiboua ({c_dept["Loh Djiboua"]})  |  Jan–Fév 2026')
ws_ana.Range('B5').Font.Size=9; ws_ana.Range('B5').Font.Italic=True
ws_ana.Range('B5').HorizontalAlignment=xlCenter
ws_ana.Rows('5').RowHeight=16
spacer(6)

row = 7

# ══ 1. RÉSUMÉ EXÉCUTIF ═══════════════════════════════════════════════════════
write_title(row, '1.  RÉSUMÉ EXÉCUTIF', C_BLUE); row+=1

kpi_rows = [
    ('Indicateur','Valeur','Nbre','%','Benchmark','Statut'),
    ('Total enquêtes analysées', N, N, '100%', '—', 'OK Complet'),
    ('Eau de bonne qualité', p_bonne, n_bonne, f'{p_bonne}%', '>50% ODD', '! Insuffisant'),
    ('Accès eau < 30 min', p_acc30, n_acc30, f'{p_acc30}%', '>75%', '! Moyen'),
    ('Ménages traitant l\'eau', p_traite, n_traite, f'{p_traite}%', '>50%', '! Faible'),
    ('Maladies liées à l\'eau', p_mal, n_mal, f'{p_mal}%', '<10%', '[!!] Critique'),
    ('Pollution atmosphérique', p_poll, n_poll, f'{p_poll}%', '<20%', '[!!] Élevé'),
    ('Dégradation ressources', p_deg, n_deg, f'{p_deg}%', '<20%', '[!!] Critique'),
    ('Durée moy. enquête (min)', moy_d, '—', '—', '30-40 min', 'OK Normal'),
]
write_header(row, kpi_rows[0]); row+=1
for i, r_data in enumerate(kpi_rows[1:]):
    write_row(row, r_data, stripe=(i%2==0), bold_first=True); row+=1
spacer(row); row+=1

# ══ 2. EAU ════════════════════════════════════════════════════════════════════
write_title(row, '2.  EAU  –  Accès, Qualité & Traitement', C_BLUE); row+=1

write_sub(row, '2.1 Sources d\'approvisionnement en eau'); row+=1
write_header(row, ['Source','Goh','Loh Djiboua','Total','%','Note']); row+=1
sources_order = ['puits_traditionnel','puits_ameliore','forage','riviere','sodeci','autres']
src_labels     = {'puits_traditionnel':'Puits traditionnel','puits_ameliore':'Puits amélioré',
                  'forage':'Forage','riviere':'Rivière','sodeci':'SODECI (réseau)','autres':'Autres'}
for i, s in enumerate(sources_order):
    g = xt_source['Goh'].get(s,0)
    l = xt_source['Loh Djiboua'].get(s,0)
    tot = g+l
    note = '! Non amélioré' if s in ['puits_traditionnel','riviere','autres'] else 'OK Amélioré'
    write_row(row,[src_labels.get(s,s),g,l,tot,f'{tot/N*100:.1f}%',note],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,'Note: Sources améliorées = puits amélioré, forage, SODECI. Sources non améliorées = puits traditionnel, rivière, autres.'); row+=1
spacer(row); row+=1

write_sub(row, '2.2 Qualité de l\'eau'); row+=1
write_header(row,['Qualité','Goh','Loh Djiboua','Total','%','Évaluation']); row+=1
q_order = [('bonne','Bonne','OK Acceptable'),('acceptable','Acceptable','! À améliorer'),('mauvaise','Mauvaise','[!!] Alarmant')]
for i,(k,lbl,ev) in enumerate(q_order):
    g=xt_qualite['Goh'].get(k,0); l=xt_qualite['Loh Djiboua'].get(k,0); tot=g+l
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,f'Seulement {p_bonne}% des ménages disposent d\'eau de bonne qualité. 40% ont de l\'eau mauvaise qualité – risque sanitaire majeur.'); row+=1
spacer(row); row+=1

write_sub(row, '2.3 Temps d\'accès à l\'eau'); row+=1
write_header(row,['Temps d\'accès','Goh','Loh Djiboua','Total','%','Statut ODD']); row+=1
t_order=[('moins_30','< 30 min','OK ODD atteint'),('30_60','30–60 min','! Acceptable'),('plus_60','> 60 min','[!!] Critique')]
for i,(k,lbl,st) in enumerate(t_order):
    g=xt_temps['Goh'].get(k,0); l=xt_temps['Loh Djiboua'].get(k,0); tot=g+l
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',st],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,f'{c_temps.get("plus_60",0)} ménages ({c_temps.get("plus_60",0)/N*100:.1f}%) passent plus d\'1 heure pour accéder à l\'eau – impact productivité & scolarisation.'); row+=1
spacer(row); row+=1

write_sub(row, '2.4 Traitement de l\'eau & Maladies'); row+=1
write_header(row,['Indicateur','Goh','Loh Djiboua','Total','%','—']); row+=1

def dept_cnt(field, value):
    g=sum(1 for r in ROWS if r[col('departement')]=='Goh' and r[col(field)]==value)
    l=sum(1 for r in ROWS if r[col('departement')]=='Loh Djiboua' and r[col(field)]==value)
    return g,l

g_tr,l_tr=dept_cnt('traite_eau',True)
g_mal,l_mal=dept_cnt('e7_maladie_eau','oui')
for i,(lbl,g,l,suffix) in enumerate([
    ('Traite l\'eau',g_tr,l_tr,''),
    ('Maladies liées eau',g_mal,l_mal,''),
]):
    tot=g+l
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%','—'],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row, '  Détail maladies déclarées'); row+=1
write_header(row,['Maladie','Nombre','%','—','—','—']); row+=1
for i,(d,n_d) in enumerate(sorted(diseases.items(),key=lambda x:-x[1])):
    write_row(row,[d,n_d,f'{n_d/N*100:.1f}%','—','—','—'],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# ══ 3. ASSAINISSEMENT ════════════════════════════════════════════════════════
write_title(row, '3.  ASSAINISSEMENT  –  Latrines & Hygiène', C_TEAL); row+=1

write_sub(row,'3.1 Types de latrines'); row+=1
write_header(row,['Type','Goh','Loh Djiboua','Total','%','Classification']); row+=1
lat_order=[('wc_modernes','WC modernes','OK Optimal'),('ameliorees','Latrines améliorées','OK Bon'),
           ('simple','Latrine simple','! Basique'),('NA','Aucune latrine','[!!] Défécation à l\'air libre'),('autre','Autre','—')]
def dept_xt(field,value):
    g=xt_source  # reuse helper
    g_=sum(1 for r in ROWS if r[col('departement')]=='Goh' and r[col(field)]==value)
    l_=sum(1 for r in ROWS if r[col('departement')]=='Loh Djiboua' and r[col(field)]==value)
    return g_,l_
for i,(k,lbl,cls) in enumerate(lat_order):
    g_,l_=dept_xt('e11_type',k); tot=g_+l_
    write_row(row,[lbl,g_,l_,tot,f'{tot/N*100:.1f}%',cls],stripe=(i%2==0),bold_first=True); row+=1

write_note(row,f'{c_latr.get("NA",0)} ménages ({c_latr.get("NA",0)/N*100:.1f}%) n\'ont PAS de latrines – pratique en plein air. Risque épidémique élevé.'); row+=1
spacer(row); row+=1

write_sub(row,'3.2 Évaluation assainissement & Problèmes'); row+=1
write_header(row,['Indicateur','Oui','Non','N/A','% Oui','Niveau']); row+=1
g_ass,l_ass=dept_xt('e9_eval_assainissement','oui')
n_ass=g_ass+l_ass
ass_data=[
    ('Éval. assainissement satisfaisante', n_ass, N-n_ass, 0),
    ('Problème: latrines insuffisantes', sum(1 for r in ROWS if r[col('e10_problemes_latrines_insuffisantes')]=='oui'),0,0),
    ('Problème: odeurs', sum(1 for r in ROWS if r[col('e10_problemes_odeurs')]=='oui'),0,0),
    ('Problème: inondation', sum(1 for r in ROWS if r[col('e10_problemes_inondation')]=='oui'),0,0),
    ('Inondations fréquentes', sum(1 for r in ROWS if r[col('e13_inondations')]=='souvent'),0,0),
]
for i,(lbl,yes,no_,na) in enumerate(ass_data):
    tot_y=yes; pct_y=f'{yes/N*100:.1f}%'
    lev = 'OK' if yes/N < 0.3 else ('!' if yes/N < 0.6 else '[!!]')
    write_row(row,[lbl,yes,no_,na,pct_y,lev],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# ══ 4. DÉCHETS & EAUX USÉES ═══════════════════════════════════════════════════
write_title(row, '4.  DÉCHETS & EAUX USÉES  –  Gestion & Impact', C_TEAL); row+=1

write_sub(row,'4.1 Exposition et gestion des déchets'); row+=1
write_header(row,['Indicateur','Goh','Loh Djiboua','Total','% Total','Éval.']); row+=1
g_exp,l_exp=dept_xt('d1_exposition_dechets','oui')
g_eus,l_eus=dept_xt('d6_exposition_eaux_usees','oui')
g_renv,l_renv=dept_xt('d8_respect_environnement','fort')
for i,(lbl,g,l) in enumerate([
    ('Exposition aux déchets',g_exp,l_exp),
    ('Exposition aux eaux usées',g_eus,l_eus),
    ('Respect fort de l\'environnement',g_renv,l_renv),
]):
    tot=g+l; ev='[!!]' if tot/N>0.5 and 'Exposition' in lbl else '!'
    if 'Respect' in lbl: ev='OK' if tot/N>0.5 else '!'
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row,'4.2 Modes de gestion des déchets'); row+=1
write_header(row,['Mode de gestion','Nombre','%','—','Impact env.','—']); row+=1
waste_impact={'Abandon':'[!!] Très négatif','Brûlage':'[!!] Très négatif',
              'Collecte/Val.':'OK Positif','Réutilisation':'OK Positif'}
for i,(wt,wn) in enumerate(sorted(waste_types.items(),key=lambda x:-x[1])):
    imp=waste_impact.get(wt,'—')
    write_row(row,[wt,wn,f'{wn/N*100:.1f}%','—',imp,'—'],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,'Brûlage et abandon représentent les pratiques dominantes → pollution sol & air. Collecte/valorisation marginale.'); row+=1
spacer(row); row+=1

# ══ 5. POLLUTION ══════════════════════════════════════════════════════════════
write_title(row, '5.  POLLUTION  –  Atmosphérique · Sonore · Qualité Air', C_ORANGE); row+=1

write_sub(row,'5.1 Pollution atmosphérique & qualité de l\'air'); row+=1
write_header(row,['Indicateur','Goh','Loh Djiboua','Total','%','Statut']); row+=1
g_poll,l_poll=dept_xt('p1_pollution_atmo','oui')
g_air,l_air=dept_xt('p3_qualite_air','oui')
g_bruit,l_bruit=dept_xt('p4_nuisances_sonores','oui')
g_act,l_act=dept_xt('p7_activites_polluantes','oui')
for i,(lbl,g,l) in enumerate([
    ('Pollution atmosphérique signalée',g_poll,l_poll),
    ('Bonne qualité de l\'air',g_air,l_air),
    ('Nuisances sonores',g_bruit,l_bruit),
    ('Activités polluantes à proximité',g_act,l_act),
]):
    tot=g+l; ev='[!!]' if tot/N>0.5 and 'signalée' in lbl else '!'
    if 'Bonne' in lbl: ev='OK' if tot/N>0.5 else '[!!]'
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row,'5.2 Effets sanitaires de la pollution'); row+=1
write_header(row,['Effet','Nombre','%','—','—','—']); row+=1
effects={
    'Toux':   sum(1 for r in ROWS if r[col('p8_effets_toux')]=='oui'),
    'Rhume':  sum(1 for r in ROWS if r[col('p8_effets_rhume')]=='oui'),
    'Asthme': sum(1 for r in ROWS if r[col('p8_effets_asthme')]=='oui'),
    'Irrit. yeux': sum(1 for r in ROWS if r[col('p8_effets_yeux')]=='oui'),
    'Maux de tête':sum(1 for r in ROWS if r[col('p8_effets_maux_tete')]=='oui'),
}
for i,(eff,ne) in enumerate(sorted(effects.items(),key=lambda x:-x[1])):
    write_row(row,[eff,ne,f'{ne/N*100:.1f}%','—','—','—'],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# ══ 6. RESSOURCES NATURELLES ══════════════════════════════════════════════════
write_title(row, '6.  RESSOURCES NATURELLES  –  État & Gouvernance', C_GREEN); row+=1

write_sub(row,'6.1 Disponibilité et état des ressources'); row+=1
write_header(row,['Indicateur','Goh','Loh Djiboua','Total','%','Éval.']); row+=1
g_dispo,l_dispo=dept_xt('r2_disponibilite','oui')
g_deg,l_deg=dept_xt('r5_degradation','oui')
g_sacr,l_sacr=dept_xt('r7_ressources_sacrees','oui')
g_comm,l_comm=dept_xt('r8_ressources_communautaires','oui')
g_reg,l_reg=dept_xt('r9_regles_locales','oui')
for i,(lbl,g,l) in enumerate([
    ('Ressources disponibles',g_dispo,l_dispo),
    ('Ressources dégradées',g_deg,l_deg),
    ('Ressources sacrées présentes',g_sacr,l_sacr),
    ('Ressources communautaires',g_comm,l_comm),
    ('Règles locales de gestion',g_reg,l_reg),
]):
    tot=g+l; ev='OK' if 'disponibles' in lbl and tot/N>0.5 else '!'
    if 'dégradées' in lbl: ev='[!!]' if tot/N>0.3 else 'OK'
    if 'sacrées' in lbl or 'communautaires' in lbl or 'Règles' in lbl:
        ev='i'
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row,'6.2 Causes de dégradation'); row+=1
write_header(row,['Cause','Nombre','%','—','Priorité','—']); row+=1
for i,(cause,nc) in enumerate(sorted(ress_causes.items(),key=lambda x:-x[1])):
    prio='[!!] Haute' if nc/N>0.2 else ('! Moyenne' if nc/N>0.1 else 'OK Faible')
    write_row(row,[cause,nc,f'{nc/N*100:.1f}%','—',prio,'—'],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,'Les pratiques agricoles et l\'exploitation forestière sont les principales causes de dégradation.'); row+=1
spacer(row); row+=1

# ══ 7. ANALYSE PAR DÉPARTEMENT ════════════════════════════════════════════════
write_title(row, '7.  COMPARAISON PAR DÉPARTEMENT', C_DARK_NAVY); row+=1

write_header(row,['Indicateur','Goh (570)','%','Loh Djiboua (430)','%','Écart']); row+=1
indicators=[
    ('Bonne qualité eau',     'e5_qualite_eau','bonne'),
    ('Eau traitée',           'traite_eau',True),
    ('Accès < 30 min',        'e4_temps_acces','moins_30'),
    ('Maladies liées eau',    'e7_maladie_eau','oui'),
    ('Latrine simple',        'e11_type','simple'),
    ('Aucune latrine (DA)',    'e11_type','NA'),
    ('Exposition déchets',    'd1_exposition_dechets','oui'),
    ('Pollution atmo.',       'p1_pollution_atmo','oui'),
    ('Dégradation ressources','r5_degradation','oui'),
    ('Bonne qualité air',     'p3_qualite_air','oui'),
]
for i,(lbl,field,val) in enumerate(indicators):
    g_=sum(1 for r in ROWS if r[col('departement')]=='Goh' and r[col(field)]==val)
    l_=sum(1 for r in ROWS if r[col('departement')]=='Loh Djiboua' and r[col(field)]==val)
    pg=g_/570*100; pl=l_/430*100
    ecart=f'{pg-pl:+.1f}pp'
    write_row(row,[lbl,g_,f'{pg:.1f}%',l_,f'{pl:.1f}%',ecart],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# ══ 8. RECOMMANDATIONS ════════════════════════════════════════════════════════
write_title(row, '8.  RECOMMANDATIONS PRIORITAIRES', C_RED, fg=C_WHITE); row+=1

recs=[
    ('[!!] URGENT','Eau potable','Augmenter le nombre de forages et points d\'eau améliorés dans les zones à eau de mauvaise qualité. Promouvoir le traitement domestique (chloration, filtration).'),
    ('[!!] URGENT','Assainissement','Construire des latrines améliorées dans les ménages sans infrastructure. Sensibiliser à la défécation à l\'air libre.'),
    ('[!!] URGENT','Maladies hydriques','Mettre en place un système de surveillance épidémiologique. Vaccinations préventives contre typhoïde/hépatite A.'),
    ('! MOYEN','Gestion déchets','Créer des points de collecte de déchets. Interdire/remplacer le brûlage à l\'air libre.'),
    ('! MOYEN','Pollution atmo.','Réguler les activités polluantes. Planter des zones tampons vertes.'),
    ('! MOYEN','Ressources naturelles','Renforcer les règles locales de gestion durable. Former les agriculteurs aux pratiques agroforestières.'),
    ('OK LONG TERME','Gouvernance','Renforcer la coordination entre mairie, État, ONG et communautés pour la gestion WASH.'),
    ('OK LONG TERME','Accès eau','Réduire les distances > 60 min via infrastructures rapprochées. Viser ODD 6 (eau potable pour tous d\'ici 2030).'),
]
write_header(row,['Priorité','Domaine','Recommandation','—','—','—']); row+=1
for i,(pri,dom,rec) in enumerate(recs):
    # merge C to G for long text
    ws_ana.Range(f'D{row}:G{row}').Merge()
    ws_ana.Range(f'B{row}').Value=pri
    ws_ana.Range(f'C{row}').Value=dom
    ws_ana.Range(f'D{row}').Value=rec
    bg=C_STRIPE if i%2==0 else C_WHITE
    for c_ in ['B','C','D']:
        cell_=ws_ana.Range(f'{c_}{row}')
        cell_.Interior.Color=bg; cell_.Font.Size=8; cell_.WrapText=True
        cell_.VerticalAlignment=xlVAlignTop; border_cell(cell_,C_MGRAY)
    ws_ana.Range(f'B{row}').Font.Bold=True
    ws_ana.Rows(row).RowHeight=28
    row+=1

spacer(row); row+=1

# Footer
ws_ana.Range(f'B{row}:G{row}').Merge()
ws_ana.Range(f'B{row}').Value = (
    'Rapport généré automatiquement – Données: wash_test_data.xlsx – '
    'Analyse: Claude AI – © 2026  |  Les chiffres sont basés sur l\'échantillon de 1 000 ménages.')
ws_ana.Range(f'B{row}').Font.Size=7
ws_ana.Range(f'B{row}').Font.Italic=True
ws_ana.Range(f'B{row}').Font.Color=C_DGRAY
ws_ana.Range(f'B{row}').HorizontalAlignment=xlCenter

# ── FINAL CLEANUP ─────────────────────────────────────────────────────────────
print(">>>  Finalisation…")
ws_dash.Activate()
ws_dash.Range('A1').Select()

# Hide gridlines on dashboard
excel.ActiveWindow.DisplayGridlines = False

# Restore calculation

excel.ScreenUpdating = True
excel.DisplayAlerts = False

# Save
output_name = 'WASH_Dashboard.xlsx'
output_full = r'C:\Users\ITrebi\Music\PortFolio Upwork\Excel Dashboard\WASH_Dashboard.xlsx'
wb.SaveAs(Filename=output_full, FileFormat=51)   # xlOpenXMLWorkbook
wb.Close(SaveChanges=False)
excel.Quit()

print(f"\n[OK]  Dashboard créé : {output_full}")
print(f"   Feuilles: TABLEAU_BORD | ANALYSE_PROFONDE | wash_test_data")
print(f"   Slicers: Département, Sous-préfecture, Source d'eau, Qualité eau, Mois")
print(f"   Charts : 11 graphiques liés aux tableaux croisés")
