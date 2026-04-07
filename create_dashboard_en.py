#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WASH Excel Dashboard Creator - Professional Dashboard with Pivot Tables, Slicers & Deep Analysis
English version
"""

import win32com.client as win32
import openpyxl
from collections import Counter, defaultdict
import os, sys, shutil, time

FILE_PATH  = r'C:\Users\ITrebi\Music\PortFolio Upwork\Excel Dashboard\wash_test_data.xlsx'
OUTPUT_PATH = r'C:\Users\ITrebi\Music\PortFolio Upwork\Excel Dashboard\WASH_Dashboard_EN.xlsx'

# -- Color helpers -------------------------------------------------------------
def rgb(r, g, b): return r + g*256 + b*65536

C_DARK_NAVY  = rgb(31,  56, 100)   # #1F3864
C_BLUE       = rgb( 0,  84, 166)   # #0054A6  - WASH brand blue
C_MED_BLUE   = rgb(46, 117, 182)   # #2E75B6
C_LIGHT_BLUE = rgb(189,215, 238)   # #BDD7EE
C_TEAL       = rgb( 0, 176, 185)   # #00B0B9
C_GREEN      = rgb(112,173,  71)   # #70AD47
C_ORANGE     = rgb(237,125,  49)   # #ED7D31
C_RED        = rgb(192,  0,   0)   # #C00000
C_AMBER      = rgb(255,192,   0)   # #FFC000
C_WHITE      = rgb(255,255, 255)
C_LGRAY      = rgb(242,242, 242)
C_MGRAY      = rgb(208,208, 208)
C_DGRAY      = rgb( 89, 89,  89)
C_BLACK      = rgb(  0,  0,   0)
C_STRIPE     = rgb(217,232, 248)   # light blue stripe

# -- Excel constants -----------------------------------------------------------
xlDatabase        = 1
xlRowField        = 1
xlColumnField     = 2
xlDataField       = 4
xlCount           = -4112
xlAverage         = -4106

xlColumnClustered = 51
xlBarClustered    = 57
xlPie             = 5
xlDoughnut        = -4120
xlColumnStacked100= 59
xlBarStacked100   = 70
xlBarStacked      = 52
xlColumnStacked   = 52

xlCenter          = -4108
xlLeft            = -4131
xlRight           = -4152
xlVAlignCenter    = -4108
xlVAlignTop       = -4160
xlVAlignBottom    = -4107

xlContinuous      = 1
xlThin            = 2
xlMedium          = -4138
xlThick           = 4
xlEdgeLeft        = 7
xlEdgeTop         = 8
xlEdgeBottom      = 9
xlEdgeRight       = 10
xlInsideVertical  = 11
xlInsideHorizontal= 12
xlAround          = 4

xlSheetHidden     = 0
xlSheetVisible    = -1
xlSheetVeryHidden = 2

xlPasteValues     = -4163
xlPasteSpecialOperationNone = -4142

# -- Pre-compute statistics ----------------------------------------------------
print(">>>  Loading data...")
wb_tmp = openpyxl.load_workbook(FILE_PATH, read_only=True)
ws_tmp = wb_tmp.active
HEADERS = [c.value for c in next(ws_tmp.iter_rows(min_row=1, max_row=1))]
ROWS    = list(ws_tmp.iter_rows(min_row=2, values_only=True))
wb_tmp.close()

N = len(ROWS)   # 1000

def col(name):    return HEADERS.index(name)
def vals(name):   return [r[col(name)] for r in ROWS]
def cnt(name):    return Counter(vals(name))
def pct(name, v): n = sum(1 for r in ROWS if r[col(name)]==v); return n, round(n/N*100,1)

# KPI values
n_bonne, p_bonne   = pct('e5_qualite_eau', 'bonne')
n_acc30, p_acc30   = pct('e4_temps_acces', 'moins_30')
n_traite,p_traite  = pct('traite_eau', True)
n_mal,   p_mal     = pct('e7_maladie_eau','oui')
n_poll,  p_poll    = pct('p1_pollution_atmo','oui')
n_deg,   p_deg     = pct('r5_degradation','oui')
durees = [r[col('duree_enquete_min')] for r in ROWS if r[col('duree_enquete_min')]]
moy_d  = round(sum(durees)/len(durees),1)

# Distributions
c_source  = cnt('e1_source_eau')
c_qualite = cnt('e5_qualite_eau')
c_temps   = cnt('e4_temps_acces')
c_latr    = cnt('e11_type')
c_mal     = cnt('e7_maladie_eau')
c_dechets = cnt('d3_satisfaction')
c_poll    = cnt('p1_pollution_atmo')
c_air     = cnt('p3_qualite_air')
c_ress    = cnt('r5_degradation')
c_dept    = cnt('departement')
c_evass   = cnt('e9_eval_assainissement')
c_bruit   = cnt('p4_nuisances_sonores')

# Months from ISO date
dates_raw = vals('date_soumission')
months_map = {'01':'January','02':'February','03':'March','04':'April',
              '05':'May','06':'June','07':'July','08':'August',
              '09':'September','10':'October','11':'November','12':'December'}
def parse_month(d):
    if not d or len(d)<7: return 'N/A'
    return months_map.get(d[5:7], d[5:7])
c_mois = Counter(parse_month(d) for d in dates_raw)

# Cross-tabs (dept x field)
def xtab(field):
    d = defaultdict(Counter)
    for r in ROWS:
        dept = r[col('departement')] or 'N/A'
        v    = r[col(field)] or 'N/A'
        d[dept][v] += 1
    return d

xt_source  = xtab('e1_source_eau')
xt_qualite = xtab('e5_qualite_eau')
xt_temps   = xtab('e4_temps_acces')
xt_mal     = xtab('e7_maladie_eau')
xt_poll    = xtab('p1_pollution_atmo')
xt_ress    = xtab('r5_degradation')

# Sub-prefectures
sous_prefs = sorted(set(r[col('sous_prefecture')] for r in ROWS if r[col('sous_prefecture')]))

# Disease details
diseases = {
    'Diarrhea':   sum(1 for r in ROWS if r[col('e7_maladies_diarrhee')]=='oui'),
    'Typhoid':    sum(1 for r in ROWS if r[col('e7_maladies_typhoide')]=='oui'),
    'Hepatitis':  sum(1 for r in ROWS if r[col('e7_maladies_hepatite')]=='oui'),
    'Cholera':    sum(1 for r in ROWS if r[col('e7_maladies_cholera')]=='oui'),
}

# Waste management breakdown
waste_types = {
    'Dumping':       sum(1 for r in ROWS if r[col('d2_gestion_dechets_abandon')]=='oui'),
    'Open burning':  sum(1 for r in ROWS if r[col('d2_gestion_dechets_brulage')]=='oui'),
    'Collection/Val.':sum(1 for r in ROWS if r[col('d2_gestion_dechets_collecte_valorisation')]=='oui'),
    'Reuse':         sum(1 for r in ROWS if r[col('d2_gestion_dechets_reutilisation')]=='oui'),
}

# Resource degradation causes
ress_causes = {
    'Agri. practices':  sum(1 for r in ROWS if r[col('r6_causes_Pratiques_agricoles')]=='oui'),
    'Forest. exploit.': sum(1 for r in ROWS if r[col('r6_causes_Exploitation_forestiere')]=='oui'),
    'Urbanization':     sum(1 for r in ROWS if r[col('r6_causes_Urbanisation')]=='oui'),
    'Mining exploit.':  sum(1 for r in ROWS if r[col('r6_causes_Exploitation_miniere')]=='oui'),
}

print(f"   {N} surveys | Good water: {p_bonne}% | Pollution: {p_poll}%")

# -- Copy source file to output ------------------------------------------------
print(">>>  Copying source file...")
shutil.copy2(FILE_PATH, OUTPUT_PATH)

# -- Open with win32com --------------------------------------------------------
print(">>>  Opening Excel...")
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = False
excel.DisplayAlerts = False
excel.ScreenUpdating = False

wb = excel.Workbooks.Open(OUTPUT_PATH)
ws_data = wb.Worksheets('wash_test_data')

# -- Add "mois" column to data (col 167 = FK) ----------------------------------
print(">>>  Adding month column...")
last_row = ws_data.UsedRange.Rows.Count  # 1001
ws_data.Cells(1, 167).Value = 'mois'
mois_vals = [parse_month(r[col('date_soumission')]) for r in ROWS]
for i, mv in enumerate(mois_vals):
    ws_data.Cells(i+2, 167).Value = mv

# -- Delete existing dashboard sheets if any -----------------------------------
for sname in ['DASHBOARD','DEEP_ANALYSIS','_PIVOTS','TABLEAU_BORD','ANALYSE_PROFONDE']:
    try:
        wb.Worksheets(sname).Delete()
    except: pass

# -- Create sheets -------------------------------------------------------------
print(">>>  Creating sheets...")
ws_piv    = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
ws_piv.Name = '_PIVOTS'

ws_dash   = wb.Sheets.Add(Before=wb.Sheets(1))
ws_dash.Name = 'DASHBOARD'

ws_ana    = wb.Sheets.Add(After=ws_dash)
ws_ana.Name = 'DEEP_ANALYSIS'

# Color tabs
ws_dash.Tab.Color = C_BLUE
ws_ana.Tab.Color  = C_TEAL

# -- Pivot Cache & Pivot Tables ------------------------------------------------
print(">>>  Creating pivot tables...")
data_addr = f"wash_test_data!R1C1:R{last_row}C167"
pc = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=data_addr)

def make_pt(dest_cell, name, row_fld, col_fld=None, val_fld='id_soumission'):
    pt = pc.CreatePivotTable(TableDestination=dest_cell, TableName=name)
    pt.PivotFields(row_fld).Orientation = xlRowField
    pt.PivotFields(row_fld).Position = 1
    if col_fld:
        pt.PivotFields(col_fld).Orientation = xlColumnField
        pt.PivotFields(col_fld).Position = 1
    pt.AddDataField(pt.PivotFields(val_fld), 'Nb', -4112)
    pt.ShowTableStyleRowStripes = True
    return pt

r = 1
pt_src  = make_pt(ws_piv.Cells(r,1),  'PT_Source',    'e1_source_eau',   'departement'); r+=12
pt_qlt  = make_pt(ws_piv.Cells(r,1),  'PT_Qualite',   'e5_qualite_eau',  'departement'); r+=8
pt_tps  = make_pt(ws_piv.Cells(r,1),  'PT_Temps',     'e4_temps_acces',  'departement'); r+=8
pt_lat  = make_pt(ws_piv.Cells(r,1),  'PT_Latrines',  'e11_type',        'departement'); r+=10
pt_mal  = make_pt(ws_piv.Cells(r,1),  'PT_Maladies',  'e7_maladie_eau',  'departement'); r+=8
pt_dch  = make_pt(ws_piv.Cells(r,1),  'PT_Dechets',   'd3_satisfaction', 'departement'); r+=8
pt_pll  = make_pt(ws_piv.Cells(r,1),  'PT_Pollution', 'p1_pollution_atmo','departement'); r+=8
pt_air  = make_pt(ws_piv.Cells(r,1),  'PT_Air',       'p3_qualite_air',  'departement'); r+=8
pt_res  = make_pt(ws_piv.Cells(r,1),  'PT_Ressources','r5_degradation',  'departement'); r+=8
pt_mois = make_pt(ws_piv.Cells(r,1),  'PT_Mois',      'mois');                           r+=15
pt_brt  = make_pt(ws_piv.Cells(r,1),  'PT_Bruit',     'p4_nuisances_sonores','departement'); r+=8
pt_evass= make_pt(ws_piv.Cells(r,1),  'PT_Evass',     'e9_eval_assainissement','departement'); r+=8

ws_piv.Visible = xlSheetVeryHidden

# -- Slicers -------------------------------------------------------------------
print(">>>  Creating slicers...")
ALL_PTS = [pt_src, pt_qlt, pt_tps, pt_lat, pt_mal, pt_dch, pt_pll, pt_air, pt_res, pt_brt, pt_evass]

def add_slicer_cache(base_pt, field, sc_name, other_pts):
    sc = wb.SlicerCaches.Add2(base_pt, field, sc_name)
    for pt in other_pts:
        try: sc.PivotTables.AddPivotTable(pt)
        except: pass
    return sc

sc_dept  = add_slicer_cache(pt_src, 'departement',     'SC_Dept',   ALL_PTS[1:])
sc_sp    = add_slicer_cache(pt_src, 'sous_prefecture', 'SC_SP',     ALL_PTS[1:])
sc_src   = add_slicer_cache(pt_src, 'e1_source_eau',   'SC_Src',    ALL_PTS[1:])
sc_qlt   = add_slicer_cache(pt_qlt, 'e5_qualite_eau',  'SC_Qlt',    [pt_src,pt_tps,pt_lat,pt_mal,pt_dch,pt_pll,pt_air,pt_res,pt_brt,pt_evass])
sc_mois_c= add_slicer_cache(pt_mois,'mois',             'SC_Mois',   ALL_PTS)

# -- ============================================================================
# -- FORMAT DASHBOARD SHEET
# -- ============================================================================
print(">>>  Formatting dashboard sheet...")

# Freeze panes - keep header visible
ws_dash.Activate()
ws_dash.Application.ActiveWindow.ActiveSheet.Cells(11,2).Select()
try:
    ws_dash.Application.ActiveWindow.FreezePanes = True
except: pass

# Set sheet background - entire used area
ws_dash.Cells.Interior.Color = C_LGRAY

# Column widths (in char units)
cwidths = {'A':1.5,'B':17,'C':17,'D':17,'E':17,'F':17,'G':17,
           'H':1.5,'I':20,'J':16,'K':1.5,'L':20,'M':1.5}
for c,w in cwidths.items():
    ws_dash.Columns(c).ColumnWidth = w

# -- HEADER BANNER (rows 1-5) --------------------------------------------------
for r in range(1,6):
    ws_dash.Rows(r).RowHeight = 16

ws_dash.Range('B1:G5').Merge()
h = ws_dash.Range('B1')
h.Value = 'DASHBOARD  \u00b7  WASH SURVEY'
h.Font.Name = 'Calibri'
h.Font.Size = 22
h.Font.Bold = True
h.Font.Color = C_WHITE
h.Interior.Color = C_BLUE
h.HorizontalAlignment = xlCenter
h.VerticalAlignment   = xlVAlignCenter

ws_dash.Range('I1:J2').Merge()
sub = ws_dash.Range('I1')
sub.Value = '~  Water  \u2022  Sanitation  \u2022  Hygiene'
sub.Font.Size = 11
sub.Font.Bold = True
sub.Font.Color = C_WHITE
sub.Interior.Color = C_TEAL
sub.HorizontalAlignment = xlCenter
sub.VerticalAlignment   = xlVAlignCenter

ws_dash.Range('I3:J3').Merge()
meta = ws_dash.Range('I3')
meta.Value = f'{N} surveys  |  2 departments  |  8 sub-prefectures'
meta.Font.Size = 8
meta.Font.Color = C_DGRAY
meta.Interior.Color = C_LGRAY
meta.HorizontalAlignment = xlCenter

ws_dash.Range('I4:J5').Merge()
ws_dash.Range('I4').Value = "C\u00f4te d'Ivoire \u2014 Jan/Feb 2026"
ws_dash.Range('I4').Font.Size = 8
ws_dash.Range('I4').Font.Italic = True
ws_dash.Range('I4').Font.Color = C_DGRAY
ws_dash.Range('I4').Interior.Color = C_LGRAY
ws_dash.Range('I4').HorizontalAlignment = xlCenter
ws_dash.Range('I4').VerticalAlignment   = xlVAlignCenter

# -- SLICERS PANEL HEADER ------------------------------------------------------
ws_dash.Range('L1:L5').Merge()
sp = ws_dash.Range('L1')
sp.Value = '~ FILTERS'
sp.Font.Size = 12
sp.Font.Bold = True
sp.Font.Color = C_WHITE
sp.Interior.Color = C_DARK_NAVY
sp.HorizontalAlignment = xlCenter
sp.VerticalAlignment   = xlVAlignCenter

# -- KPI CARDS (rows 6-9) ------------------------------------------------------
ws_dash.Rows('6:6').RowHeight = 28
ws_dash.Rows('7:7').RowHeight = 36
ws_dash.Rows('8:8').RowHeight = 18
ws_dash.Rows('9:9').RowHeight = 8   # spacer

kpi_def = [
    ('B','Total Surveys',       f'{N}',           '1,000 surveyed households', C_BLUE),
    ('C','Good Water Quality',  f'{p_bonne}%',    f'{n_bonne} surveys',        C_GREEN),
    ('D','Access < 30 min',     f'{p_acc30}%',    f'{n_acc30} surveys',        C_TEAL),
    ('E','Waterborne Diseases', f'{p_mal}%',      f'{n_mal} surveys',          C_RED),
    ('F','Atmo. Pollution',     f'{p_poll}%',     f'{n_poll} surveys',         C_AMBER),
    ('G','Resource Degradation',f'{p_deg}%',      f'{n_deg} surveys',          C_ORANGE),
]

def border_cell(cell, color, weight=xlThin):
    for edge in [xlEdgeLeft,xlEdgeTop,xlEdgeBottom,xlEdgeRight]:
        b = cell.Borders(edge)
        b.LineStyle = xlContinuous
        b.Weight = weight
        b.Color = color

for c, title, val, sub_, color in kpi_def:
    # Title row
    t = ws_dash.Range(f'{c}6')
    t.Value = title; t.Font.Size=9; t.Font.Bold=True
    t.Font.Color=C_WHITE; t.Interior.Color=color
    t.HorizontalAlignment=xlCenter; t.VerticalAlignment=xlVAlignCenter
    border_cell(t, color, xlMedium)
    # Value row
    v = ws_dash.Range(f'{c}7')
    v.Value=val; v.Font.Size=20; v.Font.Bold=True
    v.Font.Color=color; v.Interior.Color=C_WHITE
    v.HorizontalAlignment=xlCenter; v.VerticalAlignment=xlVAlignCenter
    border_cell(v, color, xlMedium)
    # Subtitle row
    s = ws_dash.Range(f'{c}8')
    s.Value=sub_; s.Font.Size=7; s.Font.Color=C_DGRAY
    s.Interior.Color=C_LGRAY
    s.HorizontalAlignment=xlCenter
    border_cell(s, color)

# -- SECTION LABELS ------------------------------------------------------------
def section_label(row_num, text, color):
    ws_dash.Rows(row_num).RowHeight = 20
    r = ws_dash.Range(f'B{row_num}:G{row_num}')
    r.Merge()
    r.Value = text
    r.Font.Size = 10; r.Font.Bold = True; r.Font.Color = C_WHITE
    r.Interior.Color = color
    r.HorizontalAlignment = xlCenter; r.VerticalAlignment = xlVAlignCenter

section_label(10, '~  WATER  \u2013  Sources \u00b7 Quality \u00b7 Access',               C_BLUE)
section_label(32, '~  SANITATION & HEALTH  \u2013  Latrines \u00b7 Diseases',             C_MED_BLUE)
section_label(54, '~  WASTE & POLLUTION  \u2013  Management \u00b7 Air \u00b7 Noise',     C_TEAL)
section_label(76, '~  NATURAL RESOURCES  \u2013  Availability \u00b7 Degradation',        C_GREEN)

# -- CHART ROW HEIGHTS ---------------------------------------------------------
for r in range(11,32): ws_dash.Rows(r).RowHeight = 10
for r in range(33,54): ws_dash.Rows(r).RowHeight = 10
for r in range(55,76): ws_dash.Rows(r).RowHeight = 10
for r in range(77,98): ws_dash.Rows(r).RowHeight = 10

# -- CHART HELPER --------------------------------------------------------------
def add_chart(ws, pt_range, ctype, title, left, top, w, h,
              legend_pos=-4107, color_list=None):
    co = ws.ChartObjects().Add(Left=left, Top=top, Width=w, Height=h)
    ch = co.Chart
    ch.SetSourceData(Source=pt_range)
    ch.ChartType = ctype
    ch.HasTitle = True
    ch.ChartTitle.Text = title
    ch.ChartTitle.Font.Size = 10
    ch.ChartTitle.Font.Bold = True
    ch.ChartTitle.Font.Color = C_DARK_NAVY
    ch.HasLegend = True
    ch.Legend.Position = legend_pos
    ch.Legend.Font.Size = 7
    ch.PlotArea.Interior.Color = C_WHITE
    ch.ChartArea.Interior.Color = C_LGRAY
    ch.ChartArea.Border.LineStyle = xlContinuous
    ch.ChartArea.Border.Color = C_MGRAY
    # Apply colors to series if provided
    if color_list:
        try:
            for i, c_ in enumerate(color_list):
                if i < ch.SeriesCollection().Count:
                    ch.SeriesCollection(i+1).Interior.Color = c_
                    ch.SeriesCollection(i+1).Format.Fill.ForeColor.RGB = c_
        except: pass
    return ch

# -- GET CELL POSITIONS --------------------------------------------------------
def pos(row, col_letter):
    c = ws_dash.Range(f'{col_letter}{row}')
    return c.Left, c.Top

# Chart dimensions
CW3 = 245   # width for 3-per-row
CW2 = 375   # width for 2-per-row
CH  = 185   # height

GAP = 4

# Row 1 of charts: rows 11-31 -> top of row 11
L_B, T10 = pos(11, 'B')
L_C = ws_dash.Range('C11').Left
L_E = ws_dash.Range('E11').Left

# -- SECTION WATER CHARTS ------------------------------------------------------
# Chart 1: Water Sources (Donut)
add_chart(ws_dash, pt_src.TableRange1,
          xlDoughnut, "Water Sources", L_B, T10, CW3, CH,
          color_list=[C_BLUE,C_TEAL,C_GREEN,C_AMBER,C_ORANGE,C_RED])

# Chart 2: Water Quality by Department (Bar Clustered)
L_C2 = L_B + CW3 + GAP
add_chart(ws_dash, pt_qlt.TableRange1,
          xlBarClustered, "Water Quality by Department",
          L_C2, T10, CW3, CH,
          color_list=[C_GREEN,C_AMBER,C_RED])

# Chart 3: Water Access Time (Column Clustered)
L_C3 = L_C2 + CW3 + GAP
add_chart(ws_dash, pt_tps.TableRange1,
          xlColumnClustered, "Water Access Time",
          L_C3, T10, CW3, CH,
          color_list=[C_GREEN,C_AMBER,C_RED])

# -- SECTION SANITATION CHARTS -------------------------------------------------
_, T32 = pos(33, 'B')

# Chart 4: Types of Latrines (Bar)
add_chart(ws_dash, pt_lat.TableRange1,
          xlBarClustered, "Types of Latrines",
          L_B, T32, CW3, CH,
          color_list=[C_BLUE,C_MED_BLUE,C_LIGHT_BLUE,C_TEAL])

# Chart 5: Waterborne Diseases (Bar)
add_chart(ws_dash, pt_mal.TableRange1,
          xlColumnClustered, "Waterborne Diseases",
          L_C2, T32, CW3, CH,
          color_list=[C_RED, C_ORANGE, C_AMBER])

# Chart 6: Sanitation Evaluation (Column)
add_chart(ws_dash, pt_evass.TableRange1,
          xlColumnClustered, "Sanitation Evaluation",
          L_C3, T32, CW3, CH,
          color_list=[C_GREEN, C_RED, C_AMBER])

# -- SECTION WASTE & POLLUTION CHARTS ------------------------------------------
_, T54 = pos(55, 'B')

# Chart 7: Waste Mgmt Satisfaction (Bar)
add_chart(ws_dash, pt_dch.TableRange1,
          xlBarClustered, "Waste Mgmt Satisfaction",
          L_B, T54, CW3, CH,
          color_list=[C_GREEN, C_RED])

# Chart 8: Atmospheric Pollution (Column)
add_chart(ws_dash, pt_pll.TableRange1,
          xlColumnClustered, "Atmospheric Pollution",
          L_C2, T54, CW3, CH,
          color_list=[C_ORANGE, C_GREEN])

# Chart 9: Air Quality (Bar)
add_chart(ws_dash, pt_air.TableRange1,
          xlBarClustered, "Air Quality",
          L_C3, T54, CW3, CH,
          color_list=[C_GREEN, C_RED])

# -- SECTION NATURAL RESOURCES CHARTS ------------------------------------------
_, T76 = pos(77, 'B')

# Chart 10: Natural Resources Degradation (Column)
add_chart(ws_dash, pt_res.TableRange1,
          xlColumnClustered, "Natural Resources Degradation",
          L_B, T76, CW2, CH,
          color_list=[C_RED, C_GREEN, C_AMBER])

# Chart 11: Noise Nuisances (Bar)
L_R2 = L_B + CW2 + GAP
add_chart(ws_dash, pt_brt.TableRange1,
          xlBarClustered, "Noise Nuisances",
          L_R2, T76, CW2, CH,
          color_list=[C_AMBER, C_GREEN])

# -- ADD SLICERS TO DASHBOARD --------------------------------------------------
print(">>>  Positioning slicers...")
L_slicer = ws_dash.Range('I6').Left
T_slicer = ws_dash.Range('I6').Top
SW = ws_dash.Range('I6:J6').Width   # slicer width
SH_S = 120  # short height
SH_M = 150  # medium height
SH_L = 180  # long height

slicer_defs = [
    (sc_dept,  'Department',    SW, SH_S),
    (sc_sp,    'Sub-prefecture',SW, SH_L),
    (sc_src,   'Water Source',  SW, SH_M),
    (sc_qlt,   'Water Quality', SW, SH_S),
    (sc_mois_c,'Month',         SW, SH_S),
]

y_off = T_slicer
for sc, cap, sw, sh in slicer_defs:
    try:
        slc = sc.Slicers.Add(ws_dash, Caption=cap,
                             Left=L_slicer, Top=y_off, Width=sw, Height=sh)
        slc.Style = 'SlicerStyleLight2'
        y_off += sh + 6
    except Exception as e:
        print(f"   Slicer '{cap}': {e}")

# -- ============================================================================
# -- DEEP_ANALYSIS SHEET
# -- ============================================================================
print(">>>  Creating deep analysis sheet...")

ws_ana.Cells.Interior.Color = C_WHITE
ws_ana.Columns('A').ColumnWidth = 3
ws_ana.Columns('B').ColumnWidth = 40
ws_ana.Columns('C').ColumnWidth = 18
ws_ana.Columns('D').ColumnWidth = 18
ws_ana.Columns('E').ColumnWidth = 18
ws_ana.Columns('F').ColumnWidth = 18
ws_ana.Columns('G').ColumnWidth = 18

def cell(row, col_letter):
    return ws_ana.Range(f'{col_letter}{row}')

def write_title(row, text, bg=C_BLUE, fg=C_WHITE, size=14, merge_to='G'):
    r = ws_ana.Range(f'B{row}:{merge_to}{row}')
    r.Merge(); r.Value=text
    r.Font.Size=size; r.Font.Bold=True
    r.Font.Color=fg; r.Interior.Color=bg
    r.HorizontalAlignment=xlCenter; r.VerticalAlignment=xlVAlignCenter
    ws_ana.Rows(row).RowHeight=24

def write_sub(row, text, bg=C_MED_BLUE, fg=C_WHITE):
    r = ws_ana.Range(f'B{row}:G{row}')
    r.Merge(); r.Value=text
    r.Font.Size=10; r.Font.Bold=True
    r.Font.Color=fg; r.Interior.Color=bg
    r.HorizontalAlignment=xlLeft; r.VerticalAlignment=xlVAlignCenter
    ws_ana.Rows(row).RowHeight=20

def write_header(row, labels, bg=C_DARK_NAVY, fg=C_WHITE):
    cols = ['B','C','D','E','F','G']
    ws_ana.Rows(row).RowHeight = 18
    for i, lbl in enumerate(labels):
        if i < len(cols):
            c = ws_ana.Range(f'{cols[i]}{row}')
            c.Value=lbl; c.Font.Bold=True; c.Font.Size=9
            c.Font.Color=fg; c.Interior.Color=bg
            c.HorizontalAlignment=xlCenter; c.VerticalAlignment=xlVAlignCenter
            border_cell(c, bg)

def write_row(row, vals, stripe=False, bold_first=False):
    cols = ['B','C','D','E','F','G']
    bg = C_STRIPE if stripe else C_WHITE
    ws_ana.Rows(row).RowHeight = 16
    for i, v in enumerate(vals):
        if i < len(cols):
            c = ws_ana.Range(f'{cols[i]}{row}')
            c.Value = v; c.Font.Size=9
            c.Interior.Color = bg
            c.VerticalAlignment = xlVAlignCenter
            if i==0 and bold_first: c.Font.Bold=True
            if i > 0: c.HorizontalAlignment = xlCenter
            border_cell(c, C_MGRAY)

def write_note(row, text, color=C_DGRAY):
    ws_ana.Range(f'B{row}:G{row}').Merge()
    c = ws_ana.Range(f'B{row}')
    c.Value = text; c.Font.Size=8; c.Font.Italic=True; c.Font.Color=color
    c.HorizontalAlignment=xlLeft
    ws_ana.Rows(row).RowHeight = 14

def spacer(row, h=8):
    ws_ana.Rows(row).RowHeight = h

# == MAIN TITLE ================================================================
ws_ana.Rows('1:1').RowHeight = 8
ws_ana.Range('B2:G4').Merge()
t = ws_ana.Range('B2')
t.Value='WASH DATA IN-DEPTH ANALYSIS'
t.Font.Size=18; t.Font.Bold=True; t.Font.Color=C_WHITE
t.Interior.Color=C_BLUE
t.HorizontalAlignment=xlCenter; t.VerticalAlignment=xlVAlignCenter
ws_ana.Rows('2:4').RowHeight=22

ws_ana.Range('B5:G5').Merge()
ws_ana.Range('B5').Value = (
    f"WASH Survey \u2013 C\u00f4te d'Ivoire  |  {N} households  |  Depts: Goh ({c_dept['Goh']}) & Loh Djiboua ({c_dept['Loh Djiboua']})  |  Jan\u2013Feb 2026")
ws_ana.Range('B5').Font.Size=9; ws_ana.Range('B5').Font.Italic=True
ws_ana.Range('B5').HorizontalAlignment=xlCenter
ws_ana.Rows('5').RowHeight=16
spacer(6)

row = 7

# == 1. EXECUTIVE SUMMARY =====================================================
write_title(row, '1.  EXECUTIVE SUMMARY', C_BLUE); row+=1

kpi_rows = [
    ('Indicator','Value','Count','%','Benchmark','Status'),
    ('Total surveyed households', N, N, '100%', '\u2014', 'OK Complete'),
    ('Good quality water', p_bonne, n_bonne, f'{p_bonne}%', '>50% SDG', '! Insufficient'),
    ('Water access < 30 min', p_acc30, n_acc30, f'{p_acc30}%', '>75%', '! Average'),
    ('Households treating water', p_traite, n_traite, f'{p_traite}%', '>50%', '! Low'),
    ('Waterborne diseases', p_mal, n_mal, f'{p_mal}%', '<10%', '[!!] Critical'),
    ('Atmospheric pollution', p_poll, n_poll, f'{p_poll}%', '<20%', '[!!] High'),
    ('Resource degradation', p_deg, n_deg, f'{p_deg}%', '<20%', '[!!] Critical'),
    ('Average survey duration (min)', moy_d, '\u2014', '\u2014', '30-40 min', 'OK Normal'),
]
write_header(row, kpi_rows[0]); row+=1
for i, r_data in enumerate(kpi_rows[1:]):
    write_row(row, r_data, stripe=(i%2==0), bold_first=True); row+=1
spacer(row); row+=1

# == 2. WATER =================================================================
write_title(row, '2.  WATER  \u2013  Access, Quality & Treatment', C_BLUE); row+=1

write_sub(row, '2.1 Water supply sources'); row+=1
write_header(row, ['Source','Goh','Loh Djiboua','Total','%','Type']); row+=1
sources_order = ['puits_traditionnel','puits_ameliore','forage','riviere','sodeci','autres']
src_labels     = {'puits_traditionnel':'Traditional well','puits_ameliore':'Improved well',
                  'forage':'Borehole','riviere':'River','sodeci':'SODECI (piped)','autres':'Other'}
for i, s in enumerate(sources_order):
    g = xt_source['Goh'].get(s,0)
    l = xt_source['Loh Djiboua'].get(s,0)
    tot = g+l
    note = '! Unimproved' if s in ['puits_traditionnel','riviere','autres'] else 'OK Improved'
    write_row(row,[src_labels.get(s,s),g,l,tot,f'{tot/N*100:.1f}%',note],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,'Note: Improved sources = improved well, borehole, SODECI. Unimproved = traditional well, river, other.'); row+=1
spacer(row); row+=1

write_sub(row, '2.2 Water quality'); row+=1
write_header(row,['Quality','Goh','Loh Djiboua','Total','%','Assessment']); row+=1
q_order = [('bonne','Good','OK Acceptable'),('acceptable','Acceptable','! To improve'),('mauvaise','Poor','[!!] Alarming')]
for i,(k,lbl,ev) in enumerate(q_order):
    g=xt_qualite['Goh'].get(k,0); l=xt_qualite['Loh Djiboua'].get(k,0); tot=g+l
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,f'Only {p_bonne}% of households have good quality water. ~40% have poor quality water \u2013 major health risk.'); row+=1
spacer(row); row+=1

write_sub(row, '2.3 Water access time'); row+=1
write_header(row,['Access time','Goh','Loh Djiboua','Total','%','SDG Status']); row+=1
t_order=[('moins_30','< 30 min','OK SDG met'),('30_60','30-60 min','! Acceptable'),('plus_60','> 60 min','[!!] Critical')]
for i,(k,lbl,st) in enumerate(t_order):
    g=xt_temps['Goh'].get(k,0); l=xt_temps['Loh Djiboua'].get(k,0); tot=g+l
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',st],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,f'{c_temps.get("plus_60",0)} households ({c_temps.get("plus_60",0)/N*100:.1f}%) spend more than 1 hour accessing water \u2013 impact on productivity & schooling.'); row+=1
spacer(row); row+=1

write_sub(row, '2.4 Water treatment & Diseases'); row+=1
write_header(row,['Indicator','Goh','Loh Djiboua','Total','%','\u2014']); row+=1

def dept_cnt(field, value):
    g=sum(1 for r in ROWS if r[col('departement')]=='Goh' and r[col(field)]==value)
    l=sum(1 for r in ROWS if r[col('departement')]=='Loh Djiboua' and r[col(field)]==value)
    return g,l

g_tr,l_tr=dept_cnt('traite_eau',True)
g_mal,l_mal=dept_cnt('e7_maladie_eau','oui')
for i,(lbl,g,l,suffix) in enumerate([
    ('Treats water',g_tr,l_tr,''),
    ('Waterborne diseases',g_mal,l_mal,''),
]):
    tot=g+l
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%','\u2014'],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row, '  Declared disease breakdown'); row+=1
write_header(row,['Disease','Count','%','\u2014','\u2014','\u2014']); row+=1
for i,(d,n_d) in enumerate(sorted(diseases.items(),key=lambda x:-x[1])):
    write_row(row,[d,n_d,f'{n_d/N*100:.1f}%','\u2014','\u2014','\u2014'],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# == 3. SANITATION ============================================================
write_title(row, '3.  SANITATION  \u2013  Latrines & Hygiene', C_TEAL); row+=1

write_sub(row,'3.1 Types of latrines'); row+=1
write_header(row,['Type','Goh','Loh Djiboua','Total','%','Classification']); row+=1
lat_order=[('wc_modernes','Modern toilets','OK Optimal'),('ameliorees','Improved latrines','OK Good'),
           ('simple','Simple latrine','! Basic'),('NA','No latrine (open def.)','[!!] Open Defecation'),('autre','Other','\u2014')]
def dept_xt(field,value):
    g_=sum(1 for r in ROWS if r[col('departement')]=='Goh' and r[col(field)]==value)
    l_=sum(1 for r in ROWS if r[col('departement')]=='Loh Djiboua' and r[col(field)]==value)
    return g_,l_
for i,(k,lbl,cls) in enumerate(lat_order):
    g_,l_=dept_xt('e11_type',k); tot=g_+l_
    write_row(row,[lbl,g_,l_,tot,f'{tot/N*100:.1f}%',cls],stripe=(i%2==0),bold_first=True); row+=1

write_note(row,f'{c_latr.get("NA",0)} households ({c_latr.get("NA",0)/N*100:.1f}%) have NO latrine \u2013 open defecation practice. High epidemic risk.'); row+=1
spacer(row); row+=1

write_sub(row,'3.2 Sanitation evaluation & Problems'); row+=1
write_header(row,['Indicator','Yes','No','N/A','% Yes','Level']); row+=1
g_ass,l_ass=dept_xt('e9_eval_assainissement','oui')
n_ass=g_ass+l_ass
ass_data=[
    ('Satisfactory sanitation eval.', n_ass, N-n_ass, 0),
    ('Problem: insufficient latrines', sum(1 for r in ROWS if r[col('e10_problemes_latrines_insuffisantes')]=='oui'),0,0),
    ('Problem: odors', sum(1 for r in ROWS if r[col('e10_problemes_odeurs')]=='oui'),0,0),
    ('Problem: flooding', sum(1 for r in ROWS if r[col('e10_problemes_inondation')]=='oui'),0,0),
    ('Frequent flooding', sum(1 for r in ROWS if r[col('e13_inondations')]=='souvent'),0,0),
]
for i,(lbl,yes,no_,na) in enumerate(ass_data):
    tot_y=yes; pct_y=f'{yes/N*100:.1f}%'
    lev = 'OK' if yes/N < 0.3 else ('!' if yes/N < 0.6 else '[!!]')
    write_row(row,[lbl,yes,no_,na,pct_y,lev],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# == 4. WASTE & WASTEWATER ====================================================
write_title(row, '4.  WASTE & WASTEWATER  \u2013  Management & Impact', C_TEAL); row+=1

write_sub(row,'4.1 Waste exposure and management'); row+=1
write_header(row,['Indicator','Goh','Loh Djiboua','Total','% Total','Eval.']); row+=1
g_exp,l_exp=dept_xt('d1_exposition_dechets','oui')
g_eus,l_eus=dept_xt('d6_exposition_eaux_usees','oui')
g_renv,l_renv=dept_xt('d8_respect_environnement','fort')
for i,(lbl,g,l) in enumerate([
    ('Waste exposure',g_exp,l_exp),
    ('Wastewater exposure',g_eus,l_eus),
    ('Strong environmental respect',g_renv,l_renv),
]):
    tot=g+l; ev='[!!]' if tot/N>0.5 and 'exposure' in lbl.lower() else '!'
    if 'respect' in lbl.lower(): ev='OK' if tot/N>0.5 else '!'
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row,'4.2 Waste management methods'); row+=1
write_header(row,['Management method','Count','%','\u2014','Env. impact','\u2014']); row+=1
waste_impact={'Dumping':'[!!] Very negative','Open burning':'[!!] Very negative',
              'Collection/Val.':'OK Positive','Reuse':'OK Positive'}
for i,(wt,wn) in enumerate(sorted(waste_types.items(),key=lambda x:-x[1])):
    imp=waste_impact.get(wt,'\u2014')
    write_row(row,[wt,wn,f'{wn/N*100:.1f}%','\u2014',imp,'\u2014'],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,'Open burning and dumping are the dominant practices -> soil & air pollution. Collection/valorization marginal.'); row+=1
spacer(row); row+=1

# == 5. POLLUTION =============================================================
write_title(row, '5.  POLLUTION  \u2013  Atmospheric \u00b7 Noise \u00b7 Air Quality', C_ORANGE); row+=1

write_sub(row,'5.1 Atmospheric pollution & air quality'); row+=1
write_header(row,['Indicator','Goh','Loh Djiboua','Total','%','Status']); row+=1
g_poll,l_poll=dept_xt('p1_pollution_atmo','oui')
g_air,l_air=dept_xt('p3_qualite_air','oui')
g_bruit,l_bruit=dept_xt('p4_nuisances_sonores','oui')
g_act,l_act=dept_xt('p7_activites_polluantes','oui')
for i,(lbl,g,l) in enumerate([
    ('Atmospheric pollution reported',g_poll,l_poll),
    ('Good air quality',g_air,l_air),
    ('Noise nuisances',g_bruit,l_bruit),
    ('Polluting activities nearby',g_act,l_act),
]):
    tot=g+l; ev='[!!]' if tot/N>0.5 and 'reported' in lbl.lower() else '!'
    if 'Good' in lbl: ev='OK' if tot/N>0.5 else '[!!]'
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row,'5.2 Health effects of pollution'); row+=1
write_header(row,['Effect','Count','%','\u2014','\u2014','\u2014']); row+=1
effects={
    'Cough':       sum(1 for r in ROWS if r[col('p8_effets_toux')]=='oui'),
    'Cold':        sum(1 for r in ROWS if r[col('p8_effets_rhume')]=='oui'),
    'Asthma':      sum(1 for r in ROWS if r[col('p8_effets_asthme')]=='oui'),
    'Eye irritation': sum(1 for r in ROWS if r[col('p8_effets_yeux')]=='oui'),
    'Headaches':   sum(1 for r in ROWS if r[col('p8_effets_maux_tete')]=='oui'),
}
for i,(eff,ne) in enumerate(sorted(effects.items(),key=lambda x:-x[1])):
    write_row(row,[eff,ne,f'{ne/N*100:.1f}%','\u2014','\u2014','\u2014'],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# == 6. NATURAL RESOURCES =====================================================
write_title(row, '6.  NATURAL RESOURCES  \u2013  Status & Governance', C_GREEN); row+=1

write_sub(row,'6.1 Resource availability and status'); row+=1
write_header(row,['Indicator','Goh','Loh Djiboua','Total','%','Eval.']); row+=1
g_dispo,l_dispo=dept_xt('r2_disponibilite','oui')
g_deg,l_deg=dept_xt('r5_degradation','oui')
g_sacr,l_sacr=dept_xt('r7_ressources_sacrees','oui')
g_comm,l_comm=dept_xt('r8_ressources_communautaires','oui')
g_reg,l_reg=dept_xt('r9_regles_locales','oui')
for i,(lbl,g,l) in enumerate([
    ('Resources available',g_dispo,l_dispo),
    ('Resources degraded',g_deg,l_deg),
    ('Sacred resources present',g_sacr,l_sacr),
    ('Community resources',g_comm,l_comm),
    ('Local management rules',g_reg,l_reg),
]):
    tot=g+l; ev='OK' if 'available' in lbl.lower() and tot/N>0.5 else '!'
    if 'degraded' in lbl.lower(): ev='[!!]' if tot/N>0.3 else 'OK'
    if 'sacred' in lbl.lower() or 'community' in lbl.lower() or 'rules' in lbl.lower():
        ev='i'
    write_row(row,[lbl,g,l,tot,f'{tot/N*100:.1f}%',ev],stripe=(i%2==0),bold_first=True); row+=1

write_sub(row,'6.2 Degradation causes'); row+=1
write_header(row,['Cause','Count','%','\u2014','Priority','\u2014']); row+=1
for i,(cause,nc) in enumerate(sorted(ress_causes.items(),key=lambda x:-x[1])):
    prio='[!!] High' if nc/N>0.2 else ('! Medium' if nc/N>0.1 else 'OK Low')
    write_row(row,[cause,nc,f'{nc/N*100:.1f}%','\u2014',prio,'\u2014'],stripe=(i%2==0),bold_first=True); row+=1
write_note(row,'Agricultural practices and forest exploitation are the main causes of degradation.'); row+=1
spacer(row); row+=1

# == 7. DEPARTMENTAL COMPARISON ===============================================
write_title(row, '7.  DEPARTMENTAL COMPARISON', C_DARK_NAVY); row+=1

write_header(row,['Indicator','Goh (570)','%','Loh Djiboua (430)','%','Gap']); row+=1
indicators=[
    ('Good water quality',      'e5_qualite_eau','bonne'),
    ('Treated water',           'traite_eau',True),
    ('Access <30min',           'e4_temps_acces','moins_30'),
    ('Waterborne diseases',     'e7_maladie_eau','oui'),
    ('Simple latrine',          'e11_type','simple'),
    ('No latrine (OD)',         'e11_type','NA'),
    ('Waste exposure',          'd1_exposition_dechets','oui'),
    ('Atmo. pollution',         'p1_pollution_atmo','oui'),
    ('Resource degradation',    'r5_degradation','oui'),
    ('Good air quality',        'p3_qualite_air','oui'),
]
for i,(lbl,field,val) in enumerate(indicators):
    g_=sum(1 for r in ROWS if r[col('departement')]=='Goh' and r[col(field)]==val)
    l_=sum(1 for r in ROWS if r[col('departement')]=='Loh Djiboua' and r[col(field)]==val)
    pg=g_/570*100; pl=l_/430*100
    ecart=f'{pg-pl:+.1f}pp'
    write_row(row,[lbl,g_,f'{pg:.1f}%',l_,f'{pl:.1f}%',ecart],stripe=(i%2==0),bold_first=True); row+=1
spacer(row); row+=1

# == 8. PRIORITY RECOMMENDATIONS ==============================================
write_title(row, '8.  PRIORITY RECOMMENDATIONS', C_RED, fg=C_WHITE); row+=1

recs=[
    ('[!!] URGENT','Water','Increase boreholes and improved water points. Promote household water treatment (chlorination, filtration).'),
    ('[!!] URGENT','Sanitation','Build improved latrines in households without sanitation. Sensitize against open defecation.'),
    ('[!!] URGENT','Waterborne diseases','Establish epidemiological surveillance. Preventive vaccinations against typhoid/hepatitis A.'),
    ('! MEDIUM','Waste management','Create waste collection points. Ban/replace open burning.'),
    ('! MEDIUM','Atmo. pollution','Regulate polluting activities. Plant green buffer zones.'),
    ('! MEDIUM','Natural resources','Strengthen local sustainable management rules. Train farmers in agroforestry.'),
    ('OK LONG TERM','Governance','Strengthen coordination between municipality, State, NGOs and communities.'),
    ('OK LONG TERM','Water access','Reduce distances >60 min via closer infrastructure. Target SDG 6 (clean water for all by 2030).'),
]
write_header(row,['Priority','Domain','Recommendation','\u2014','\u2014','\u2014']); row+=1
for i,(pri,dom,rec) in enumerate(recs):
    # merge D to G for long text
    ws_ana.Range(f'D{row}:G{row}').Merge()
    ws_ana.Range(f'B{row}').Value=pri
    ws_ana.Range(f'C{row}').Value=dom
    ws_ana.Range(f'D{row}').Value=rec
    bg=C_STRIPE if i%2==0 else C_WHITE
    for c_ in ['B','C','D']:
        cell_=ws_ana.Range(f'{c_}{row}')
        cell_.Interior.Color=bg; cell_.Font.Size=8; cell_.WrapText=True
        cell_.VerticalAlignment=xlVAlignTop; border_cell(cell_,C_MGRAY)
    ws_ana.Range(f'B{row}').Font.Bold=True
    ws_ana.Rows(row).RowHeight=28
    row+=1

spacer(row); row+=1

# Footer
ws_ana.Range(f'B{row}:G{row}').Merge()
ws_ana.Range(f'B{row}').Value = (
    'Report auto-generated \u2013 Data: wash_test_data.xlsx \u2013 '
    'Analysis: Claude AI \u2013 2026  |  Figures based on sample of 1,000 households.')
ws_ana.Range(f'B{row}').Font.Size=7
ws_ana.Range(f'B{row}').Font.Italic=True
ws_ana.Range(f'B{row}').Font.Color=C_DGRAY
ws_ana.Range(f'B{row}').HorizontalAlignment=xlCenter

# -- FINAL CLEANUP -------------------------------------------------------------
print(">>>  Finalizing...")
ws_dash.Activate()
ws_dash.Range('A1').Select()

# Hide gridlines on dashboard
excel.ActiveWindow.DisplayGridlines = False

# Restore screen updating
excel.ScreenUpdating = True
excel.DisplayAlerts = False

# Save
output_full = r'C:\Users\ITrebi\Music\PortFolio Upwork\Excel Dashboard\WASH_Dashboard_EN.xlsx'
wb.SaveAs(Filename=output_full, FileFormat=51)   # xlOpenXMLWorkbook
wb.Close(SaveChanges=False)
excel.Quit()

print(f"\n[OK]  Dashboard created: {output_full}")
print(f"   Sheets: DASHBOARD | DEEP_ANALYSIS | wash_test_data")
print(f"   Slicers: Department, Sub-prefecture, Water Source, Water Quality, Month")
print(f"   Charts : 11 charts linked to pivot tables")
